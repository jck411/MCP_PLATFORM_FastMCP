"""
FastMCP Desktop Example

A comprehensive example demonstrating resource and tool patterns for handling
large, structured, and binary files efficiently.

NOTESwap out the bridge tools for native resources/read calls once your model 
provider supports them—your index resource and chunk helpers stay the same.
"""

from __future__ import annotations

import itertools
import json
import mimetypes
from pathlib import Path
from typing import Any, List

from mcp.server.fastmcp import FastMCP

# Create server
mcp = FastMCP("Demo")

# Base directory for resources
RESOURCE_DIR = Path("/home/jack/Documents/MCP.resources")

@mcp.resource(
    "resource://docs-index",
    name="DocumentIndex",
    description="Index of available documents with metadata",
    mime_type="application/json",
)
def docs_index() -> List[dict[str, Any]]:
    """Return list of available documents with basic metadata."""
    return [
        {
            "name": f.name,
            "size": f.stat().st_size,
            "mime": mimetypes.guess_type(f)[0] or "application/octet-stream",
            "uri": str(f.absolute())
        }
        for f in RESOURCE_DIR.iterdir() if f.is_file()
    ]

@mcp.tool()
def read_text_chunk(
    uri: str,
    offset: int = 0,
    length: int = 8192,
) -> str:
    """Read a slice of a text file.
    
    Args:
        uri: Path to the text file
        offset: Starting byte offset
        length: Maximum number of bytes to read (default: 8KB)
        
    Returns:
        The requested text chunk
    """
    # Guard against excessive reads
    if length > 2 * 1024 * 1024:  # 2MB limit
        raise ValueError("Requested chunk size exceeds 2MB limit")
        
    with open(uri, "r", encoding="utf-8") as fp:
        fp.seek(offset)
        return fp.read(length)

@mcp.tool()
def extract_pdf_pages(
    uri: str,
    pages: List[int],
) -> List[str]:
    """Extract text from specific PDF pages.
    
    Args:
        uri: Path to the PDF file
        pages: List of page numbers (0-based) to extract
        
    Returns:
        List of extracted text strings, one per page
    """
    import fitz  # PyMuPDF
    
    # Guard against excessive page requests
    if len(pages) > 50:
        raise ValueError("Cannot extract more than 50 pages at once")
        
    with fitz.open(uri) as doc:
        max_page = len(doc) - 1
        valid_pages = [p for p in pages if 0 <= p <= max_page]
        return [doc.load_page(p).get_text() for p in valid_pages]

@mcp.tool()
def get_sheet_rows(
    uri: str,
    sheet: str,
    start: int,
    n: int = 50,
) -> List[List[Any]]:
    """Fetch rows from an Excel sheet.
    
    Args:
        uri: Path to the Excel file
        sheet: Name of the sheet to read
        start: Starting row number (0-based)
        n: Number of rows to read (default: 50)
        
    Returns:
        List of rows, where each row is a list of cell values
    """
    import openpyxl
    
    # Guard against excessive row requests
    if n > 200:
        raise ValueError("Cannot fetch more than 200 rows at once")
        
    wb = openpyxl.load_workbook(uri, read_only=True)
    ws = wb[sheet]
    rows = itertools.islice(ws.iter_rows(values_only=True), start, start + n)
    return json.loads(json.dumps(list(rows)))  # Ensure JSON-serializable values

@mcp.tool()
def get_sheet_schema(uri: str) -> dict[str, Any]:
    """Get metadata about an Excel workbook.
    
    Args:
        uri: Path to the Excel file
        
    Returns:
        Dictionary containing sheet names and dimensions
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(uri, read_only=True)
    return {
        "sheets": [
            {
                "name": name,
                "max_row": wb[name].max_row,
                "max_column": wb[name].max_column
            }
            for name in wb.sheetnames
        ]
    }

if __name__ == "__main__":
    mcp.run()